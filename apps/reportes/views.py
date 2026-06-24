import io
from collections import defaultdict

from django.contrib.auth.decorators import login_required
from django.db.models import Count
from django.db.models.functions import TruncMonth
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from django.utils import timezone

from apps.clasificacion.models import Residuo, ResultadoAdministracion


# ─── Dashboard principal ─────────────────────────────────────────────────────

@login_required
def dashboard(request):
    total      = Residuo.objects.count()
    analizados = ResultadoAdministracion.objects.count()
    pendientes = total - analizados
    alertas    = sum(
        len(r["alertas"]) for r in ResultadoAdministracion.objects.values("alertas")
        if r["alertas"]
    )

    return render(request, 'reportes/dashboard.html', {
        'total':      total,
        'analizados': analizados,
        'pendientes': pendientes,
        'alertas':    alertas,
    })


# ─── APIs JSON para Chart.js ─────────────────────────────────────────────────

@login_required
def api_materiales(request):
    data = defaultdict(int)
    for row in ResultadoAdministracion.objects.values("material_predominante").annotate(total=Count("id")):
        data[row["material_predominante"]] = row["total"]

    colores = {
        'Metal':        '#6b7280',
        'Vidrio':       '#0ea5e9',
        'Plástico':     '#3b82f6',
        'Papel/Cartón': '#f59e0b',
        'Orgánico':     '#22c55e',
        'Espuma/EPS':   '#e5e7eb',
        'Otro':         '#1f2937',
        'Sin clasificar': '#d1d5db',
    }

    labels  = list(data.keys())
    valores = list(data.values())
    colors  = [colores.get(l, '#94a3b8') for l in labels]

    return JsonResponse({'labels': labels, 'data': valores, 'colors': colors})


@login_required
def api_temporal(request):
    qs = (
        Residuo.objects
        .annotate(mes=TruncMonth('fecha'))
        .values('mes')
        .annotate(total=Count('id'))
        .order_by('mes')
    )
    labels = [r['mes'].strftime('%b %Y') for r in qs]
    data   = [r['total'] for r in qs]
    return JsonResponse({'labels': labels, 'data': data})


@login_required
def api_prioridades(request):
    qs = (
        ResultadoAdministracion.objects
        .values('prioridad')
        .annotate(total=Count('id'))
        .order_by('prioridad')
    )
    colores = {'Alta': '#ef4444', 'Media': '#f59e0b', 'Baja': '#22c55e'}
    labels  = [r['prioridad'] for r in qs]
    data    = [r['total']     for r in qs]
    colors  = [colores.get(l, '#94a3b8') for l in labels]
    return JsonResponse({'labels': labels, 'data': data, 'colors': colors})


@login_required
def api_confianza(request):
    """Distribución de confianza por rangos (0-25%, 25-50%, 50-75%, 75-100%)."""
    rangos  = {'0–25%': 0, '25–50%': 0, '50–75%': 0, '75–100%': 0}
    for row in Residuo.objects.values_list('confianza', flat=True):
        c = (row or 0) * 100
        if   c < 25:  rangos['0–25%']   += 1
        elif c < 50:  rangos['25–50%']  += 1
        elif c < 75:  rangos['50–75%']  += 1
        else:         rangos['75–100%'] += 1
    return JsonResponse({'labels': list(rangos.keys()), 'data': list(rangos.values())})


# ─── Exportar Excel ──────────────────────────────────────────────────────────

@login_required
def exportar_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse('openpyxl no está instalado. Ejecuta: pip install openpyxl',
                            status=500)

    wb = openpyxl.Workbook()

    # ── Hoja 1: Residuos ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Residuos'

    header_fill = PatternFill(start_color='1b4332', end_color='1b4332', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    thin = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    headers1 = ['ID', 'Fecha captura', 'Tipo objeto', 'Categoría',
                 'Confianza', 'Material predominante', 'Prioridad',
                 'Nivel prioridad', 'Alertas', 'Fecha procesado']

    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin

    ws1.row_dimensions[1].height = 22

    qs = ResultadoAdministracion.objects.select_related('residuo').order_by(
        'nivel_prioridad', '-fecha_procesado')

    for row_idx, a in enumerate(qs, 2):
        row_data = [
            a.residuo.id,
            a.residuo.fecha.strftime('%d/%m/%Y %H:%M'),
            a.residuo.tipo,
            a.residuo.categoria,
            round(a.residuo.confianza, 3),
            a.material_predominante,
            a.prioridad,
            a.nivel_prioridad,
            '; '.join(a.alertas) if a.alertas else '',
            a.fecha_procesado.strftime('%d/%m/%Y %H:%M') if a.fecha_procesado else '',
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col, value=val)
            cell.border = thin
            if col in (1, 5, 8):
                cell.alignment = center_align

    # Ajustar ancho de columnas
    col_widths = [6, 18, 24, 16, 10, 20, 10, 14, 50, 18]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── Hoja 2: Materiales ────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Materiales')

    material_data = defaultdict(int)
    for row in ResultadoAdministracion.objects.values("material_predominante").annotate(total=Count("id")):
        material_data[row["material_predominante"]] = row["total"]

    ws2.cell(1, 1, 'Material').font       = header_font
    ws2.cell(1, 1).fill                   = header_fill
    ws2.cell(1, 2, 'Cantidad').font       = header_font
    ws2.cell(1, 2).fill                   = header_fill
    ws2.cell(1, 3, 'Porcentaje').font     = header_font
    ws2.cell(1, 3).fill                   = header_fill
    total_mat = sum(material_data.values()) or 1

    for i, (mat, cnt) in enumerate(sorted(material_data.items(), key=lambda x: -x[1]), 2):
        ws2.cell(i, 1, mat)
        ws2.cell(i, 2, cnt)
        ws2.cell(i, 3, f'{cnt/total_mat*100:.1f}%')

    for col_letter in ['A', 'B', 'C']:
        ws2.column_dimensions[col_letter].width = 18

    # ── Generar respuesta ─────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nombre = f'reporte_ecotrack_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre}"'

    try:
        from apps.historial.utils import registrar
        registrar(request, 'EXPORTAR', 'reportes', f'Exportación Excel generada: {nombre}')
    except Exception:
        pass

    return response


# ─── Imprimir reporte (HTML para PDF via navegador) ──────────────────────────

@login_required
def imprimir_reporte(request):
    total      = Residuo.objects.count()
    analizados = ResultadoAdministracion.objects.count()
    pendientes = total - analizados

    material_data = defaultdict(int)
    for row in ResultadoAdministracion.objects.values("material_predominante").annotate(total=Count("id")):
        material_data[row["material_predominante"]] = row["total"]

    registros = list(
        ResultadoAdministracion.objects
        .select_related('residuo')
        .order_by('nivel_prioridad', '-fecha_procesado')[:50]
    )

    try:
        from apps.historial.utils import registrar
        registrar(request, 'EXPORTAR', 'reportes', 'Reporte PDF generado.')
    except Exception:
        pass

    return render(request, 'reportes/imprimir.html', {
        'total': total, 'analizados': analizados, 'pendientes': pendientes,
        'material_data': dict(material_data),
        'registros': registros,
        'fecha': timezone.now(),
    })
